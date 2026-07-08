"""
清关文件复核系统 — Customs Document Review System
基于 Streamlit + pdfplumber + PaddleOCR + DeepSeek API
"""
import streamlit as st
import pandas as pd
import tempfile
import os
import json
import hashlib
import re
from pathlib import Path
from datetime import datetime
from io import BytesIO

# ── 页面配置 ─────────────────────────────────────────────
st.set_page_config(
    page_title="清关文件复核系统",
    page_icon="🛃",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ── 样式 ─────────────────────────────────────────────────
st.markdown("""
<style>
    .match-item { background-color: #d4edda; padding: 8px 12px; border-radius: 6px; margin: 4px 0; border-left: 4px solid #28a745; }
    .mismatch-item { background-color: #f8d7da; padding: 8px 12px; border-radius: 6px; margin: 4px 0; border-left: 4px solid #dc3545; }
    .info-item { background-color: #d1ecf1; padding: 8px 12px; border-radius: 6px; margin: 4px 0; border-left: 4px solid #17a2b8; }
    .block-container { padding-top: 2rem; }
</style>
""", unsafe_allow_html=True)

# ── 侧边栏配置 ──────────────────────────────────────────
with st.sidebar:
    st.image("https://img.icons8.com/fluency/96/customs.png", width=64)
    st.title("🛃 清关文件复核")

    st.divider()
    st.subheader("⚙️ API 配置")

    # DeepSeek API Key — 优先从 Streamlit Secrets 读取
    default_key = ""
    try:
        default_key = st.secrets["DEEPSEEK_API_KEY"]
    except Exception:
        default_key = os.environ.get("DEEPSEEK_API_KEY", "")

    api_key = st.text_input(
        "DeepSeek API Key",
        type="password",
        value=default_key,
        help="从 platform.deepseek.com 获取。Streamlit Cloud 部署时已配置 Secrets 则自动填入。",
        placeholder="sk-...",
    )

    # 模型选择
    model = st.selectbox(
        "比对模型",
        options=["deepseek-chat", "deepseek-reasoner"],
        index=0,
        format_func=lambda x: "DeepSeek-V3 (推荐)" if x == "deepseek-chat" else "DeepSeek-R1 (深度推理)",
        help="DeepSeek-V3 速度快成本低；DeepSeek-R1 推理更深但较慢。",
    )

    st.divider()
    st.subheader("📋 使用说明")
    st.markdown("""
    1. **上传标准文件** — 作为比对基准的 PDF 清关文件
    2. **上传待复核文件** — 需要核对的 PDF 或图片
    3. **开始复核** — AI 自动提取并比对两份文件的数据
    4. **查看结果** — 一致/不一致项一目了然，支持导出 Excel
    """)

    st.divider()
    st.caption("Powered by DeepSeek · pdfplumber · PaddleOCR")

# ── 文件提取函数 ─────────────────────────────────────────

@st.cache_data(show_spinner=False, ttl=3600)
def extract_text_from_pdf(file_bytes: bytes, filename: str) -> str:
    """用 pdfplumber 提取 PDF 文本和表格，返回合并后的结构化文本。"""
    import pdfplumber

    full_text = []
    try:
        with pdfplumber.open(BytesIO(file_bytes)) as pdf:
            for i, page in enumerate(pdf.pages, 1):
                # 提取文本
                text = page.extract_text()
                if text:
                    full_text.append(f"--- 第 {i} 页 ---\n{text}")

                # 提取表格
                tables = page.extract_tables()
                for j, table in enumerate(tables, 1):
                    if table:
                        full_text.append(f"\n[表格 {i}.{j}]")
                        for row in table:
                            row_text = " | ".join(str(c) if c else "" for c in row)
                            full_text.append(row_text)
    except Exception as e:
        st.warning(f"pdfplumber 提取时出现问题: {e}")
        return ""

    result = "\n".join(full_text)
    if not result.strip():
        return ""  # 可能是扫描件，返回空字符串触发 OCR
    return result


# ── Tesseract OCR ────────────────────────────────────────

def tesseract_ocr(file_bytes: bytes, filename: str) -> str:
    """用 Tesseract OCR 提取扫描件/图片中的文字。"""
    import pytesseract, fitz
    from PIL import Image

    full_text = []
    ext = Path(filename).suffix.lower()

    try:
        if ext == ".pdf":
            doc = fitz.open(stream=file_bytes, filetype="pdf")
            for i, page in enumerate(doc, 1):
                pix = page.get_pixmap(dpi=200)
                img = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)
                text = pytesseract.image_to_string(img, lang="chi_sim+eng")
                if text.strip():
                    full_text.append(f"--- 第 {i} 页 ---\n{text}")
            doc.close()
        else:
            img = Image.open(BytesIO(file_bytes))
            text = pytesseract.image_to_string(img, lang="chi_sim+eng")
            if text.strip():
                full_text.append(text)
        return "\n".join(full_text)
    except Exception as e:
        st.warning(f"Tesseract OCR 失败: {e}")
        return ""


# PaddleOCR 可用性检测
try:
    from paddleocr import PaddleOCR

    HAS_PADDLEOCR = True
except ImportError:
    HAS_PADDLEOCR = False


def extract_text_with_pymupdf(file_bytes: bytes, filename: str) -> str:
    """用 PyMuPDF 提取扫描 PDF 的嵌入文本（轻量级后备方案）。"""
    try:
        import fitz

        full_text = []
        pdf_doc = fitz.open(stream=file_bytes, filetype="pdf")
        for i, page in enumerate(pdf_doc, 1):
            text = page.get_text()
            if text.strip():
                full_text.append(f"--- 第 {i} 页 ---\n{text}")
        pdf_doc.close()
        return "\n".join(full_text) if full_text else ""
    except Exception:
        return ""


def extract_text_with_image_ocr(file_bytes: bytes, filename: str) -> str:
    """OCR 后备链：PaddleOCR → Tesseract → PyMuPDF。"""
    # PaddleOCR（本地可用时）
    if HAS_PADDLEOCR:
        try:
            import fitz
            ocr = PaddleOCR(lang="ch", use_angle_cls=True, show_log=False)
            full_text = []

            if Path(filename).suffix.lower() == ".pdf":
                doc = fitz.open(stream=file_bytes, filetype="pdf")
                for page in doc:
                    pix = page.get_pixmap(dpi=200)
                    img_bytes = pix.tobytes("png")
                    with tempfile.NamedTemporaryFile(suffix=".png", delete=False) as tmp:
                        tmp.write(img_bytes); tmp_path = tmp.name
                    try:
                        r = ocr.ocr(tmp_path)
                        if r and r[0]:
                            full_text.append("\n".join(l[1][0] for l in r[0]))
                    finally:
                        os.unlink(tmp_path)
                doc.close()
            else:
                with tempfile.NamedTemporaryFile(suffix=Path(filename).suffix, delete=False) as tmp:
                    tmp.write(file_bytes); tmp_path = tmp.name
                try:
                    r = ocr.ocr(tmp_path)
                    if r and r[0]:
                        full_text.append("\n".join(l[1][0] for l in r[0]))
                finally:
                    os.unlink(tmp_path)
            return "\n".join(full_text)
        except Exception:
            pass

    # Tesseract（Streamlit Cloud 后备）
    text = tesseract_ocr(file_bytes, filename)
    if text.strip():
        return text

    # PyMuPDF（最后尝试）
    return extract_text_with_pymupdf(file_bytes, filename)


def smart_extract(file_bytes: bytes, filename: str, api_key: str = "") -> tuple:
    """
    智能提取：pdfplumber → Tesseract OCR → 报错提示。
    返回 (extracted_text, method_used)
    api_key 参数保留用于 DeepSeek 比对环节，OCR 不需要。
    """
    ext = Path(filename).suffix.lower()

    # 图片文件：Tesseract OCR
    if ext in (".png", ".jpg", ".jpeg", ".bmp", ".tiff", ".tif"):
        st.info("正在用 Tesseract OCR 识别图片文字...")
        text = extract_text_with_image_ocr(file_bytes, filename)
        if text.strip():
            return text, "Tesseract OCR (图片)"
        return "", "OCR 失败"

    # PDF: 先试 pdfplumber
    if ext == ".pdf":
        text = extract_text_from_pdf(file_bytes, filename)

        if len(text.strip()) < 50:
            st.info("检测到扫描件 PDF，正在用 Tesseract OCR 识别...")
            ocr_text = extract_text_with_image_ocr(file_bytes, filename)
            if ocr_text and len(ocr_text.strip()) > 0:
                return ocr_text, "Tesseract OCR (扫描 PDF)"
            elif text.strip():
                return text, "pdfplumber (文本 PDF — 内容较少)"
            st.warning("⚠️ 扫描件 PDF OCR 失败。请确认文件清晰度。")
            return "", "提取失败（OCR 无结果）"
        return text, "pdfplumber (文本 PDF)"

    return "", "不支持的文件格式"


def detect_data_fields(text: str) -> dict:
    """
    从提取的文本中自动识别清关相关的数据字段。
    返回 {字段名: 值} 字典。
    """
    fields = {}

    # 常见清关字段的正则匹配模式
    patterns = {
        "报关单号": r"(?:报关单号|海关编号|Declaration\s*No)[：:\s]*([A-Za-z0-9\-]+)",
        "提单号": r"(?:提单号|B/L\s*No|Bill\s*of\s*Lading\s*No)[：:\s]*([A-Za-z0-9\-]+)",
        "合同号": r"(?:合同号|Contract\s*No)[：:\s]*([A-Za-z0-9\-]+)",
        "发票号": r"(?:发票号|Invoice\s*No)[：:\s]*([A-Za-z0-9\-]+)",
        "集装箱号": r"(?:集装箱号|箱号|Container\s*No)[：:\s]*([A-Za-z0-9\-]+)",
        "总毛重": r"(?:总毛重|毛重|Gross\s*Weight)[：:\s]*([\d.,]+\s*(?:KG|KGS|千克|公斤|kg|kgs)?)",
        "总净重": r"(?:总净重|净重|Net\s*Weight)[：:\s]*([\d.,]+\s*(?:KG|KGS|千克|公斤|kg|kgs)?)",
        "总金额": r"(?:总金额|总价|Total\s*Amount|Total\s*Value)[：:\s]*([\d.,]+\s*(?:USD|EUR|CNY|美元|人民币)?)",
        "贸易条款": r"(?:贸易条款|贸易方式|Terms?\s*of\s*Trade|Incoterms?)[：:\s]*([A-Za-z]{3,4})",
        "起运港": r"(?:起运港|装货港|Port\s*of\s*Loading)[：:\s]*([\u4e00-\u9fa5A-Za-z\s]+)",
        "目的港": r"(?:目的港|卸货港|Port\s*of\s*Discharge)[：:\s]*([\u4e00-\u9fa5A-Za-z\s]+)",
        "发货人": r"(?:发货人|Shipper|Consigner)[：:\s]*([\u4e00-\u9fa5A-Za-z\s,.()（）]+?)(?:\n|$)",
        "收货人": r"(?:收货人|Consignee)[：:\s]*([\u4e00-\u9fa5A-Za-z\s,.()（）]+?)(?:\n|$)",
        "商品名称": r"(?:商品名称|货物名称|品名|Description\s*of\s*Goods)[：:\s]*([\u4e00-\u9fa5A-Za-z\s,.\-]+?)(?:\n|型号|规格|$)",
        "HS编码": r"(?:HS\s*(?:编码|代码|Code)?|商品编码|税号)[：:\s]*([\d.]{4,12})",
        "原产国": r"(?:原产国|原产地|Country\s*of\s*Origin)[：:\s]*([\u4e00-\u9fa5A-Za-z\s]+)",
        "件数": r"(?:件数|数量|Packages?|Number\s*of\s*Packages?)[：:\s]*([\d,]+\s*(?:件|PCS|CTNS|PKGS|箱)?)",
        "包装类型": r"(?:包装类型|包装|Package\s*Type)[：:\s]*([\u4e00-\u9fa5A-Za-z\s]+)",
        "运输方式": r"(?:运输方式|运输|Mode\s*of\s*Transport)[：:\s]*([\u4e00-\u9fa5A-Za-z\s]+)",
        "船名": r"(?:船名|Vessel|Vessel\s*Name)[：:\s]*([\u4e00-\u9fa5A-Za-z0-9\s\-]+)",
        "航次": r"(?:航次|Voyage\s*No)[：:\s]*([A-Za-z0-9\-]+)",
    }

    for field, pattern in patterns.items():
        match = re.search(pattern, text, re.IGNORECASE | re.MULTILINE)
        if match:
            fields[field] = match.group(1).strip()

    return fields


# ── DeepSeek 比对 ────────────────────────────────────────

def compare_with_deepseek(
    standard_text: str,
    review_text: str,
    standard_fields: dict,
    review_fields: dict,
    api_key: str,
    model: str,
) -> dict:
    """
    调用 DeepSeek API 进行基准比对（仅核心字段：收发货人 + 单号）。
    返回结构化比对结果。
    """
    from openai import OpenAI

    client = OpenAI(api_key=api_key, base_url="https://api.deepseek.com")
    max_chars = 8000

    prompt = f"""你是清关文件审核员。请比对基准文件和待复核文件的**核心字段**是否一致。

【比对范围 — 仅以下字段】
- 发货人 / Shipper
- 收货人 / Consignee
- 合同号 / Contract No
- 发票号 / Invoice No
- 提单号 / B/L No

【基准文件字段】
{json.dumps(standard_fields, ensure_ascii=False, indent=2)}

【待复核文件字段】
{json.dumps(review_fields, ensure_ascii=False, indent=2)}

【基准文件文本（前3000字符）】
{standard_text[:3000]}

【待复核文件文本（前3000字符）】
{review_text[:3000]}

只输出 JSON：

{{
  "matches": [{{"field": "字段名", "standard_value": "基准值", "review_value": "复核值", "remark": ""}}],
  "mismatches": [{{"field": "字段名", "standard_value": "基准值", "review_value": "复核值", "severity": "high|medium|low", "detail": "差异"}}],
  "only_in_standard": [],
  "only_in_review": [],
  "summary": "结论（30字内）"
}}

规则：
1. 名称允许中英/繁简差异但实质须一致
2. 单号必须逐字符一致
3. 不一致 severity：单号不同=high，收发货人不同=medium"""

    try:
        response = client.chat.completions.create(
            model=model,
            messages=[
                {"role": "system", "content": "你是清关审核专家，只输出 JSON。"},
                {"role": "user", "content": prompt},
            ],
            temperature=0.1, max_tokens=4096,
            response_format={"type": "json_object"} if model == "deepseek-chat" else None,
        )
        result_text = response.choices[0].message.content.strip()
        json_match = re.search(r"\{[\s\S]*\}", result_text)
        if json_match:
            result_text = json_match.group(0)
        result = json.loads(result_text)

        if not result.get("matches") and not result.get("mismatches"):
            result["_warning"] = "AI 未识别到可比对的核心字段。请检查文件内容。"
            result["_raw_text_sample"] = standard_text[:500] + "\n...\n" + review_text[:500]
        return result

    except json.JSONDecodeError:
        err_text = result_text[:800] if "result_text" in dir() else "无返回内容"
        st.error(f"❌ DeepSeek 返回非 JSON。\\n```\\n{err_text}\\n```")
        return {"matches": [], "mismatches": [], "only_in_standard": [], "only_in_review": [],
                "summary": "比对失败：API 返回格式异常", "_error": f"JSON parse error"}
    except Exception as e:
        st.error(f"❌ DeepSeek API 调用失败: {e}")
        return {"matches": [], "mismatches": [], "only_in_standard": [], "only_in_review": [],
                "summary": f"比对失败: {str(e)}", "_error": str(e)}


def check_internal_consistency(text: str, api_key: str, model: str) -> dict:
    """
    检查单文件内部数据一致性（单价/金额/重量/托盘/港口）。
    """
    from openai import OpenAI

    client = OpenAI(api_key=api_key, base_url="https://api.deepseek.com")

    prompt = f"""你是清关文件审核员。请检查以下文件**内部关键数据是否自洽**。

【仅检查以下字段】
- 单品单价 / Unit Price
- 单品小计金额 / Amount（数量×单价是否≈小计金额）
- 单品毛重 / Gross Weight
- 单品净重 / Net Weight（毛重应 ≥ 净重）
- 总金额 / Total Amount（单品金额累加是否≈总金额）
- 总重量 / Total Weight（单品重量累加是否≈总重量）
- 总托盘数量 / Total Pallets
- 发货港 / Port of Loading
- 目的地港 / Port of Discharge

【文件文本（前6000字符）】
{text[:6000]}

只输出 JSON：
{{
  "consistent": true,
  "issues": [{{"field": "字段", "location": "位置描述", "detail": "不一致说明", "severity": "high|medium|low"}}],
  "summary": "内部一致性结论（30字内）"
}}

注意：OCR 可能有错行，数值接近（差异<5%）可视为一致，明显不合逻辑才报告。港口名称允许中英文差异。"""

    try:
        response = client.chat.completions.create(
            model=model,
            messages=[
                {"role": "system", "content": "你是清关审核专家，只输出 JSON。"},
                {"role": "user", "content": prompt},
            ],
            temperature=0.1, max_tokens=4096,
            response_format={"type": "json_object"} if model == "deepseek-chat" else None,
        )
        result_text = response.choices[0].message.content.strip()
        json_match = re.search(r"\{[\s\S]*\}", result_text)
        if json_match:
            result_text = json_match.group(0)
        return json.loads(result_text)
    except Exception as e:
        return {"consistent": True, "issues": [], "summary": f"一致性检查异常: {str(e)}", "_error": str(e)}


# ── Excel 导出 ───────────────────────────────────────────

def export_to_excel(result: dict, shipper_label: str, review_name: str) -> bytes:
    """将比对结果导出为 Excel 文件。"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()

    # ── 概览 Sheet ──
    ws_overview = wb.active
    ws_overview.title = "比对概览"

    title_font = Font(name="微软雅黑", size=14, bold=True, color="1F4E79")
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    ws_overview["A1"] = "清关文件复核报告"
    ws_overview["A1"].font = title_font
    ws_overview.merge_cells("A1:D1")

    ws_overview["A3"] = "核对发货人:"
    ws_overview["B3"] = shipper_label
    ws_overview["A4"] = "待复核文件:"
    ws_overview["B4"] = review_name
    ws_overview["A5"] = "复核时间:"
    ws_overview["B5"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    matches = result.get("matches", [])
    mismatches = result.get("mismatches", [])

    ws_overview["A7"] = "一致项数量:"
    ws_overview["B7"] = len(matches)
    ws_overview["A8"] = "不一致项数量:"
    ws_overview["B8"] = len(mismatches)
    ws_overview["A9"] = "一致性率:"
    total = len(matches) + len(mismatches)
    rate = f"{len(matches)/total*100:.1f}%" if total > 0 else "N/A"
    ws_overview["B9"] = rate

    ws_overview["A11"] = "复核结论:"
    ws_overview["B11"] = result.get("summary", "")

    for col in ["A", "B", "C", "D"]:
        ws_overview.column_dimensions[col].width = 20

    # ── 不一致项 Sheet ──
    ws_mismatch = wb.create_sheet("不一致项")

    headers = ["字段", "标准文件值", "待复核文件值", "严重程度", "差异说明"]
    for col_idx, header in enumerate(headers, 1):
        cell = ws_mismatch.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    severity_fills = {
        "high": PatternFill("solid", fgColor="FFC7CE"),
        "medium": PatternFill("solid", fgColor="FFEB9C"),
        "low": PatternFill("solid", fgColor="C6EFCE"),
    }

    for row_idx, item in enumerate(mismatches, 2):
        values = [
            item.get("field", ""),
            item.get("standard_value", ""),
            item.get("review_value", ""),
            item.get("severity", ""),
            item.get("detail", ""),
        ]
        fill = severity_fills.get(item.get("severity", ""), None)
        for col_idx, val in enumerate(values, 1):
            cell = ws_mismatch.cell(row=row_idx, column=col_idx, value=val)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            if fill:
                cell.fill = fill

    ws_mismatch.freeze_panes = "A2"
    col_widths = [18, 25, 25, 10, 40]
    for i, w in enumerate(col_widths, 1):
        ws_mismatch.column_dimensions[get_column_letter(i)].width = w

    # ── 一致项 Sheet ──
    ws_match = wb.create_sheet("一致项")
    match_headers = ["字段", "标准文件值", "待复核文件值", "备注"]
    for col_idx, header in enumerate(match_headers, 1):
        cell = ws_match.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    for row_idx, item in enumerate(matches, 2):
        values = [
            item.get("field", ""),
            item.get("standard_value", ""),
            item.get("review_value", ""),
            item.get("remark", ""),
        ]
        for col_idx, val in enumerate(values, 1):
            cell = ws_match.cell(row=row_idx, column=col_idx, value=val)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center", wrap_text=True)

    ws_match.freeze_panes = "A2"
    for i, w in enumerate([18, 25, 25, 40], 1):
        ws_match.column_dimensions[get_column_letter(i)].width = w

    output = BytesIO()
    wb.save(output)
    return output.getvalue()


# ── 主界面 ───────────────────────────────────────────────

st.title("🛃 清关文件复核系统")
st.markdown("填写正确收发货人，上传待复核文件，AI 自动提取并校验。")

# ── 收发货人输入 ──
st.subheader("📋 正确收发货人")
col_shipper, col_consignee = st.columns(2)
with col_shipper:
    correct_shipper = st.text_input(
        "正确发货人 (Shipper)",
        placeholder="例如：DAIT INTERNATIONAL LTD",
        help="填入正确的发货人全称，比对时不区分大小写。",
    )
with col_consignee:
    correct_consignee = st.text_input(
        "正确收货人 (Consignee)",
        placeholder="例如：ABC TRADING CO., LTD",
        help="填入正确的收货人全称，比对时不区分大小写。",
    )

# ── 待复核文件区域 ──
st.subheader("📑 待复核文件")
review_files = st.file_uploader(
    "上传一份或多份待复核文件（PDF / 图片）",
    type=["pdf", "png", "jpg", "jpeg", "bmp", "tiff", "tif"],
    key="review",
    accept_multiple_files=True,
    help="支持的格式：PDF（文本型/扫描型）、PNG、JPG、BMP、TIFF。",
)

# ── 比对按钮 ──
st.divider()
btn_col1, btn_col2, btn_col3 = st.columns([1, 1, 2])

with btn_col1:
    start_btn = st.button(
        "🔍 开始复核",
        type="primary",
        use_container_width=True,
        disabled=not (correct_shipper and correct_consignee and review_files and api_key),
    )

with btn_col2:
    clear_btn = st.button("🗑️ 清除结果", use_container_width=True)

with btn_col3:
    missing = []
    if not api_key:
        missing.append("⚙️ 请在左侧侧边栏填写 DeepSeek API Key")
    if not correct_shipper:
        missing.append("📋 请填写正确发货人")
    if not correct_consignee:
        missing.append("📋 请填写正确收货人")
    if not review_files:
        missing.append("📑 请上传待复核文件")
    if missing:
        st.warning(" · ".join(missing))

if clear_btn:
    if "results_cache" in st.session_state:
        del st.session_state["results_cache"]
    st.rerun()


def fuzzy_match(text: str, target: str) -> tuple:
    """不区分大小写、忽略空格的中英文模糊匹配。"""
    import unicodedata

    def normalize(s):
        s = unicodedata.normalize("NFKC", s).casefold()
        s = re.sub(r"\s+", "", s)
        s = s.replace("（", "(").replace("）", ")").replace("，", ",").replace("。", ".")
        return s

    t_norm = normalize(target)
    txt_norm = normalize(text)
    if t_norm in txt_norm:
        return True, "精确匹配"
    # 尝试逐词匹配（至少70%的词命中）
    t_words = set(t_norm.split(",")[0].split())
    found = sum(1 for w in t_words if w and w in txt_norm)
    ratio = found / max(len(t_words), 1)
    if ratio >= 0.7:
        return True, f"模糊匹配 ({ratio:.0%})"
    return False, "不匹配"


# ── 执行复核 ──
if start_btn:
    with st.status("正在处理...", expanded=True) as status:
        all_results = []
        for idx, review_file in enumerate(review_files):
            st.write(f"  [{idx+1}/{len(review_files)}] 正在处理: **{review_file.name}**...")
            rev_bytes = review_file.read()
            rev_text, rev_method = smart_extract(rev_bytes, review_file.name, api_key)
            rev_fields = detect_data_fields(rev_text)

            # 文本有效性检查
            if len(rev_text.strip()) < 20:
                all_results.append({
                    "file_name": review_file.name,
                    "extract_method": rev_method,
                    "extract_chars": len(rev_text),
                    "detected_fields": len(rev_fields),
                    "shipper_match": None, "consignee_match": None,
                    "matches": [], "mismatches": [], "summary": "文件提取失败",
                    "_warning": f"提取文字仅 {len(rev_text)} 字符，无法复核。",
                    "_skipped": True,
                })
                continue

            # Step 1: 收发货人比对（纯字符串，不调 API）
            shipper_ok, shipper_detail = fuzzy_match(rev_text, correct_shipper)
            consignee_ok, consignee_detail = fuzzy_match(rev_text, correct_consignee)

            matches, mismatches = [], []
            if shipper_ok:
                matches.append({"field": "发货人", "standard_value": correct_shipper,
                                "review_value": shipper_detail, "remark": ""})
            else:
                mismatches.append({"field": "发货人", "standard_value": correct_shipper,
                                   "review_value": shipper_detail, "severity": "high",
                                   "detail": f"未在文件中找到 '{correct_shipper}'"})
            if consignee_ok:
                matches.append({"field": "收货人", "standard_value": correct_consignee,
                                "review_value": consignee_detail, "remark": ""})
            else:
                mismatches.append({"field": "收货人", "standard_value": correct_consignee,
                                   "review_value": consignee_detail, "severity": "high",
                                   "detail": f"未在文件中找到 '{correct_consignee}'"})

            result = {
                "file_name": review_file.name,
                "shipper_match": shipper_ok,
                "consignee_match": consignee_ok,
                "matches": matches,
                "mismatches": mismatches,
                "summary": f"发货人{'✅' if shipper_ok else '❌'} 收货人{'✅' if consignee_ok else '❌'}",
            }

            # Step 2: 内部一致性（API 调用，仅关键财务/物流字段）
            st.write(f"  🔍 内部数据校验: **{review_file.name}**...")
            internal = check_internal_consistency(rev_text, api_key, model)
            result["internal_consistency"] = internal

            result["extract_method"] = rev_method
            result["extract_chars"] = len(rev_text)
            result["detected_fields"] = len(rev_fields)
            all_results.append(result)

        status.update(label="✅ 处理完成！", state="complete")

    st.session_state["results_cache"] = {
        "all_results": all_results,
        "correct_shipper": correct_shipper,
        "correct_consignee": correct_consignee,
    }
    st.rerun()

# ── 结果显示 ──
if "results_cache" in st.session_state:
    cache = st.session_state["results_cache"]
    all_results = cache["all_results"]
    correct_shipper = cache.get("correct_shipper", "")
    correct_consignee = cache.get("correct_consignee", "")

    st.divider()
    st.subheader("📊 复核结果")

    total_matches = sum(len(r.get("matches", [])) for r in all_results)
    total_mismatches = sum(len(r.get("mismatches", [])) for r in all_results)
    total_items = total_matches + total_mismatches

    metric_cols = st.columns(5)
    metric_cols[0].metric("复核文件数", len(all_results))
    metric_cols[1].metric("✅ 一致项", total_matches)
    metric_cols[2].metric("❌ 不一致项", total_mismatches,
                          delta=f"-{total_mismatches}" if total_mismatches > 0 else None)
    metric_cols[3].metric("一致率",
                          f"{total_matches/total_items*100:.1f}%" if total_items > 0 else "N/A")
    metric_cols[4].metric("发货人", correct_shipper[:12] + "…" if len(correct_shipper) > 12 else correct_shipper)

    # 空结果诊断
    if total_items == 0:
        st.warning("⚠️ 未产生任何比对结果。可能原因：\n\n"
                   "1. 上传的是**扫描件 PDF**（图片型），pdfplumber 提取不到文字 → 请确保文件为**文本型 PDF**\n"
                   "2. 文件内容不含清关字段（纯图片、手写等）\n"
                   "3. DeepSeek API 返回了空结果 → 检查 API Key 配额和网络")

    st.divider()

    # 逐份展示结果
    for idx, result in enumerate(all_results):
        file_name = result.get("file_name", f"文件 {idx+1}")
        matches = result.get("matches", [])
        mismatches = result.get("mismatches", [])
        summary = result.get("summary", "")

        # 诊断信息
        if result.get("_warning"):
            with st.expander(f"⚠️ {file_name} — 未识别到可比对数据"):
                st.warning(result["_warning"])
                if result.get("_raw_text_sample"):
                    st.caption("提取到的文本样本（前500字符）：")
                    st.code(result["_raw_text_sample"][:1000], language=None)
                st.caption(f"提取方式：{result.get('extract_method', 'N/A')} | "
                          f"字符数：{result.get('extract_chars', 0):,} | "
                          f"识别字段数：{result.get('detected_fields', 0)}")
        if result.get("_error"):
            st.error(f"❌ {file_name} — API 错误: {result['_error']}")

        with st.expander(
            f"{'✅' if len(mismatches) == 0 else '❌'} {file_name} — 一致 {len(matches)} / 不一致 {len(mismatches)}",
            expanded=idx == 0,
        ):
            if summary:
                st.markdown(f'<div class="info-item">📝 <b>比对结论：</b>{summary}</div>', unsafe_allow_html=True)

            if matches:
                st.markdown(f"#### ✅ 一致项 ({len(matches)})")
                cols = st.columns(3)
                for i, m in enumerate(matches):
                    with cols[i % 3]:
                        st.markdown(
                            f'<div class="match-item"><b>{m.get("field", "")}</b><br>'
                            f'标准值：{m.get("standard_value", "")}<br>'
                            f'复核值：{m.get("review_value", "")}</div>',
                            unsafe_allow_html=True,
                        )

            if mismatches:
                st.markdown(f"#### ❌ 不一致项 ({len(mismatches)})")
                for m in mismatches:
                    severity_emoji = {"high": "🔴", "medium": "🟡", "low": "🟢"}.get(
                        m.get("severity", ""), "⚪"
                    )
                    st.markdown(
                        f'<div class="mismatch-item">'
                        f'{severity_emoji} <b>{m.get("field", "")}</b> '
                        f'[{m.get("severity", "unknown").upper()} 严重]<br>'
                        f'标准值：<code>{m.get("standard_value", "")}</code><br>'
                        f'复核值：<code>{m.get("review_value", "")}</code><br>'
                        f'说明：{m.get("detail", "")}</div>',
                        unsafe_allow_html=True,
                    )

            only_std = result.get("only_in_standard", [])
            only_rev = result.get("only_in_review", [])
            if only_std or only_rev:
                cols = st.columns(2)
                if only_std:
                    cols[0].markdown("**仅在标准文件中：**")
                    for item in only_std:
                        cols[0].markdown(f"- {item}")
                if only_rev:
                    cols[1].markdown("**仅在复核文件中：**")
                    for item in only_rev:
                        cols[1].markdown(f"- {item}")

            # ── 内部一致性检查 ──
            internal = result.get("internal_consistency", {})
            if internal:
                st.markdown(f"#### 🔍 文件内部一致性")
                consistent = internal.get("consistent", True)
                issues = internal.get("issues", [])

                if not consistent:
                    st.markdown(f'<div class="info-item">⚠️ 发现 {len(issues)} 处数据矛盾</div>', unsafe_allow_html=True)
                    for iss in issues:
                        sev = {"high": "🔴", "medium": "🟡", "low": "🟢"}.get(iss.get("severity", ""), "⚪")
                        st.markdown(
                            f'<div class="mismatch-item">{sev} <b>{iss.get("field", "")}</b> '
                            f'[{iss.get("severity", "unknown").upper()}]<br>'
                            f'{iss.get("detail", "")}<br>'
                            f'<small>位置：{iss.get("location", "")}</small></div>',
                            unsafe_allow_html=True,
                        )
                elif issues:
                    st.info(f"ℹ️ 数据基本一致，{len(issues)} 处轻微提醒")
                    for iss in issues:
                        st.caption(f"· {iss.get('field', '')}: {iss.get('detail', '')}")
                else:
                    st.success(f"✅ {internal.get('summary', '文件内部数据自洽，未发现矛盾。')}")

                if internal.get("_error"):
                    st.caption(f"⚠️ {internal['_error']}")

            st.caption(
                f"提取方式：{result.get('extract_method', 'N/A')} | "
                f"提取字符数：{result.get('extract_chars', 0):,} | "
                f"识别字段数：{result.get('detected_fields', 0)}"
            )

            excel_data = export_to_excel(result, correct_shipper, file_name)
            st.download_button(
                label=f"📥 导出 {file_name} 比对报告 (Excel)",
                data=excel_data,
                file_name=f"清关复核_{Path(file_name).stem}_{datetime.now():%Y%m%d_%H%M%S}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key=f"export_{idx}",
            )

    # 批量导出
    if len(all_results) > 1:
        st.divider()
        st.markdown("### 📦 批量导出全部结果")
        from openpyxl import Workbook

        wb = Workbook()
        wb.active.title = "汇总"
        for r in all_results:
            sheet_name = Path(r.get("file_name", "unknown")).stem[:31]
            ws = wb.create_sheet(sheet_name)
            ws["A1"] = "字段"
            ws["B1"] = "比对结果"
            ws["C1"] = "标准文件值"
            ws["D1"] = "待复核文件值"
            ws["E1"] = "严重程度"
            ws["F1"] = "说明"
            row = 2
            for m in r.get("matches", []):
                ws.cell(row=row, column=1, value=m.get("field", ""))
                ws.cell(row=row, column=2, value="一致")
                ws.cell(row=row, column=3, value=m.get("standard_value", ""))
                ws.cell(row=row, column=4, value=m.get("review_value", ""))
                ws.cell(row=row, column=6, value=m.get("remark", ""))
                row += 1
            for m in r.get("mismatches", []):
                ws.cell(row=row, column=1, value=m.get("field", ""))
                ws.cell(row=row, column=2, value="不一致")
                ws.cell(row=row, column=3, value=m.get("standard_value", ""))
                ws.cell(row=row, column=4, value=m.get("review_value", ""))
                ws.cell(row=row, column=5, value=m.get("severity", ""))
                ws.cell(row=row, column=6, value=m.get("detail", ""))
                row += 1

        batch_output = BytesIO()
        wb.save(batch_output)
        st.download_button(
            label="📥 一键导出全部比对报告 (Excel)",
            data=batch_output.getvalue(),
            file_name=f"清关复核_全部_{datetime.now():%Y%m%d_%H%M%S}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="export_all",
        )

else:
    st.info("👆 上传标准文件和待复核文件后，点击「开始复核」按钮。")
